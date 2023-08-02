# -*- coding: utf-8 -*-

import os
import glob
from typing import List, Any
from openpyxl import Workbook
from openpyxl.styles import Alignment
from xmindparser import xmind_to_dict

def resolve_path(dict_, lists, title):
     """
     通过递归取出每个主分支下的所有小分支并将其作为一个列表
     :param dict_:
     :param lists:
     :param title:
     :return:
     """
     # 去除title首尾空格
     title = title.strip()
     # 若title为空，则直接取value
     if len(title) == 0:
         concat_title = dict_["title"].strip()
     else:
         concat_title = title + "\t" + dict_["title"].strip()
     if not dict_.__contains__("topics"):
         lists.append(concat_title)
     else:
         for d in dict_["topics"]:
             resolve_path(d, lists, concat_title)

def xmind_to_excel(list_, excel_path):

     f = Workbook()
     sheet = f.active

    #  # 生成单sheet的Excel文件，sheet名自取
    # #  sheet = ws.add_sheet("XX模块", cell_overwrite_ok=True)

     # 第一行固定的表头标题
     row_header = ["模块", "类型", "功能点"]
     for i in range(0, len(row_header)):
         sheet.cell(1, i+1).value=row_header[i]

     # 增量索引
     index = 0

     for h in range(0, len(list_)):
         lists: List[Any] = []
         resolve_path(list_[h], lists, "")

         for j in range(0, len(lists)):
             # 将主分支下的小分支构成列表
             lists[j] = lists[j].split('\t')
            #  print(lists[j])

             for n in range(0, len(lists[j])):
                 row = j+1 + index + 1
                 column = n + 1
                 cell = sheet.cell(row, column)
                 cell.value = lists[j][n]
                 cell.alignment = Alignment(horizontal='general', vertical='center', wrap_text=True, shrink_to_fit=True)
                 
                 # 自定义内容，比如：测试点/用例标题、预期结果、实际结果、操作步骤、优先级……
                 # 这里为了更加灵活，除序号、模块、功能点的标题固定，其余以【自定义+序号】命名，如：自定义1，需生成Excel表格后手动修改
                 if n >= 2:
                     sheet.cell(1, n+1 + 1).value=("自定义" + str(n - 1))
             # 遍历完lists并给增量索引赋值，跳出for j循环，开始for h循环
             if j == len(lists) - 1:
                 index += len(lists)

     for i in range(0, sheet.max_row):
         #设置单元格高度自适应
         sheet.row_dimensions[i+1].height = None
         
     # 设置列宽自适应
     for col in sheet.columns:
         max_length = 0
         column = col[0].column_letter  # Get the column name
        #  print("===========>")
        #  print(column)
         
         cell_rows = []
         cell_column = 0
         cell_value = None
         for cell in col:
            if cell != None and cell.value != None:
                for text in str(cell.value).split('/n'):
                    # print(text)
                    if len(text) > max_length:
                        max_length = len(text)

            # 合并同一列的重复单元格
            if cell_value == cell.value and cell.value != None:
                cell_rows.append(cell.row)

            if len(cell_rows) > 1 and (cell_value != cell.value or cell.row == len(col)):
                start = min(cell_rows)
                end = max(cell_rows)
                sheet.merge_cells(start_row=start, start_column=cell_column, end_row=end, end_column=cell_column)

            if cell_value == None or cell.value == None or cell_value != cell.value:
                cell_value = cell.value
                cell_column = cell.column
                cell_rows = []
                cell_rows.append(cell.row)

         #设置单元格宽度
         adjusted_width = max(20, max_length * 0.5)
         sheet.column_dimensions[column].width = adjusted_width
    
     print("导出地址为："+excel_path)
     f.save(excel_path)

def run(xmind_path):
     # 将XMind转化成字典
     xmind_dict = xmind_to_dict(xmind_path)
     # print("将XMind中所有内容提取出来并转换成列表：", xmind_dict)
     # Excel文件与XMind文件保存在同一目录下
     excel_name = xmind_path.split('/')[-1].split(".")[0] + '.xlsx'
     excel_path = "/".join(xmind_path.split('/')[:-1]) + "/" + excel_name
    #  print(excel_path)
     # print("通过切片得到所有分支的内容：", xmind_dict[0]['topic']['topics'])
     xmind_to_excel(xmind_dict[0]['topic']['topics'], excel_path)

def getFileName(path, suffix):
    path_list = []
    file_list = os.listdir(path)
    for i in file_list:
        if os.path.splitext(i)[1] == suffix:
            path_list.append(os.path.join(path, i))
    
    # print(path_list)
    return path_list

if __name__ == '__main__':
    # 用例地址
    path = os.getcwd()
    file_list = getFileName(path, ".xmind")

    for i in file_list:
        run(i)